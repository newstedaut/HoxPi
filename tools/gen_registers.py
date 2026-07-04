#!/usr/bin/env python3
"""Erzeugt registers.json aus der offiziellen Hoval-Modbus-Datenpunktliste (xlsx).

Die xlsx wird NICHT mitgeliefert (Urheberrecht Hoval AG). Download:
  https://www.hoval.com/misc/TTE/TTE-GW-Modbus-datapoints.xlsx
(Link steht auch in der Montageanleitung des Hoval Modbus-Gateways, Seite 14.)

Aufruf:
  python3 gen_registers.py <datenpunktliste.xlsx> [unit_ids] [ausgabe.json]

  unit_ids  Kommagetrennt, Standard "1,520,143"
            (WEZ=1, Lueftung HV=520, Puffermodul PS=143 - je Werksauslieferung;
             andere Busadressen: siehe UnitID-Tabelle in der Hoval-Anleitung)
"""
import json, sys

def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    xlsx = sys.argv[1]
    unit_ids = {int(x) for x in (sys.argv[2] if len(sys.argv) > 2 else "1,520,143").split(",")}
    out_path = sys.argv[3] if len(sys.argv) > 3 else "registers.json"

    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    sheet = next((s for s in wb.sheetnames if s.lower().startswith("eng")), wb.sheetnames[0])
    ws = wb[sheet]

    hdr, rows, seen = None, [], set()
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(h or "").strip().lower() for h in row]
            i = {h: n for n, h in enumerate(hdr)}
            c = {k: i[k] for k in ("register address", "unitname", "unitid", "functiongroup",
                                   "functionnumber", "datapointid", "datapointname",
                                   "typename", "decimal", "min. value", "max. value",
                                   "writable", "unit")}
            continue
        try:
            reg, uid = int(row[c["register address"]]), int(row[c["unitid"]])
        except (TypeError, ValueError):
            continue
        if uid not in unit_ids or reg in seen:
            continue
        seen.add(reg)
        def num(key):
            v = row[c[key]]
            try: return int(v)
            except (TypeError, ValueError): return None
        rows.append({
            "reg": reg,
            "unit_name": str(row[c["unitname"]] or ""),
            "unit_id": uid,
            "fg": num("functiongroup") or 0,
            "fn": num("functionnumber") or 0,
            "dp": num("datapointid") or 0,
            "name": str(row[c["datapointname"]] or ""),
            "type": str(row[c["typename"]] or ""),
            "decimal": num("decimal") or 0,
            "min": num("min. value"),
            "max": num("max. value"),
            "writable": str(row[c["writable"]] or "No"),
            "unit": str(row[c["unit"]] or ""),
        })
    rows.sort(key=lambda r: r["reg"])
    json.dump(rows, open(out_path, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"{out_path}: {len(rows)} Register (UnitIds {sorted(unit_ids)}, Sheet '{sheet}')")

if __name__ == "__main__":
    main()
