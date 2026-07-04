#!/usr/bin/env python3
"""Erzeugt reg_texts.json (Namen + Beschreibungen DE/EN) aus der Hoval-Datenpunktliste.

Die xlsx wird NICHT mitgeliefert (Urheberrecht Hoval AG). Download:
  https://www.hoval.com/misc/TTE/TTE-GW-Modbus-datapoints.xlsx

Aufruf:
  python3 gen_reg_texts.py <datenpunktliste.xlsx> <registers.json> [ausgabe.json]

Hinweis: Hoval pflegt englische Beschreibungen nur lueckenhaft - wo sie fehlen,
zeigt das Dashboard die deutsche Beschreibung als Fallback.
"""
import json, re, sys

def harvest(wb, sheetname, wanted):
    ws = wb[sheetname]
    hdr, out = None, {}
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(h or "").strip() for h in row]
            idx = {h.lower(): i for i, h in enumerate(hdr)}
            c_reg, c_name = idx["register address"], idx["datapointname"]
            c_com, c_type = idx["commentary"], idx["typename"]
            texts = sorted((int(re.match(r"text (\d+)$", h.lower()).group(1)), i)
                           for i, h in enumerate(hdr) if re.match(r"text \d+$", h.lower()))
            continue
        try:
            reg = int(row[c_reg])
        except (TypeError, ValueError):
            continue
        if reg not in wanted or reg in out:
            continue
        name = str(row[c_name] or "").strip()
        com = str(row[c_com] or "").strip()
        enum = []
        if str(row[c_type] or "").strip().upper() == "LIST":
            for num, i in texts:
                if row[i] not in (None, ""):
                    enum.append(f"{num}={row[i]}")
        d = com + ((" " if com else "") + "[" + ", ".join(enum) + "]" if enum else "")
        out[reg] = (name, d)
    return out

def main():
    if len(sys.argv) < 3:
        print(__doc__); sys.exit(1)
    import openpyxl
    wb = openpyxl.load_workbook(sys.argv[1], read_only=True)
    regs = json.load(open(sys.argv[2], encoding="utf-8"))
    wanted = {r["reg"] for r in regs}
    out_path = sys.argv[3] if len(sys.argv) > 3 else "reg_texts.json"

    de = harvest(wb, "Deutsch", wanted) if "Deutsch" in wb.sheetnames else {}
    en_sheet = next((s for s in wb.sheetnames if s.lower().startswith("eng")), None)
    en = harvest(wb, en_sheet, wanted) if en_sheet else {}

    texts = {}
    for reg in sorted(wanted):
        nd, dd = de.get(reg, ("", ""))
        ne, ed = en.get(reg, ("", ""))
        texts[str(reg)] = {"nd": nd, "dd": dd, "ne": ne, "ed": ed}
    json.dump(texts, open(out_path, "w", encoding="utf-8"), ensure_ascii=False, indent=0)
    n_dd = sum(1 for t in texts.values() if t["dd"])
    n_ed = sum(1 for t in texts.values() if t["ed"])
    print(f"{out_path}: {len(texts)} Register, Beschreibungen DE={n_dd}, EN={n_ed}")

if __name__ == "__main__":
    main()
